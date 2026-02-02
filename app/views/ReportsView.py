from PyQt6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QTabWidget, QTableWidget, QTableWidgetItem,
    QPushButton, QDateEdit, QLabel, QComboBox, QSpinBox, QFileDialog, QHeaderView,
    QWidget, QMessageBox
)
from PyQt6.QtCore import QDate, Qt
from datetime import datetime, timedelta
import csv
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from app.views.BaseView import BaseView
from app.utils.ProfessionalPDFReportGenerator import ProfessionalPDFReportGenerator
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.styles.numbers import FORMAT_NUMBER_00
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

class ReportsView(BaseView):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        
        
        self.tab_widget = QTabWidget()
        
        
        self.tab_widget.addTab(self.create_sales_tab(), "Sales Report")
        
    
        self.tab_widget.addTab(self.create_top_items_tab(), "Top Selling Items")
        
        
        self.tab_widget.addTab(self.create_inventory_tab(), "Inventory Status")
        
        
        self.tab_widget.addTab(self.create_low_stock_tab(), "Low Stock Alert")
        

        self.tab_widget.addTab(self.create_supplier_tab(), "Supplier Performance")
        
        self.tab_widget.addTab(self.create_category_tab(), "Category Performance")
        
        layout.addWidget(self.tab_widget)
        self.setLayout(layout)
        self.connect_export_buttons()

    def create_sales_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Date range selection
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("From:"))
        self.sales_start_date = QDateEdit()
        self.sales_start_date.setDate(QDate.currentDate().addDays(-30))
        self.sales_start_date.setCalendarPopup(True)
        date_layout.addWidget(self.sales_start_date)
        
        date_layout.addWidget(QLabel("To:"))
        self.sales_end_date = QDateEdit()
        self.sales_end_date.setDate(QDate.currentDate())
        self.sales_end_date.setCalendarPopup(True)
        date_layout.addWidget(self.sales_end_date)
        
        self.sales_refresh_btn = QPushButton("Refresh")
        self.sales_refresh_btn.setObjectName("search_button")
        date_layout.addWidget(self.sales_refresh_btn)
        
        self.sales_export_btn = QPushButton("Export to CSV")
        self.sales_export_btn.setObjectName("secondary_button")
        date_layout.addWidget(self.sales_export_btn)
        self.sales_export_pdf_btn = QPushButton("Export to PDF")
        self.sales_export_pdf_btn.setObjectName("secondary_button")
        date_layout.addWidget(self.sales_export_pdf_btn)
        
        date_layout.addStretch()
        layout.addLayout(date_layout)
        
        # Sales by date table
        layout.addWidget(QLabel("Sales Summary by Date:"))
        self.sales_by_date_table = QTableWidget()
        self.sales_by_date_table.setColumnCount(4)
        self.sales_by_date_table.setHorizontalHeaderLabels(
            ["Date", "Transactions", "Total Revenue", "Avg Transaction"]
        )
        self.sales_by_date_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.sales_by_date_table)
        
        # Detailed transactions table
        layout.addWidget(QLabel("Detailed Transactions:"))
        self.sales_detail_table = QTableWidget()
        self.sales_detail_table.setColumnCount(6)
        self.sales_detail_table.setHorizontalHeaderLabels(
            ["Transaction ID", "Date", "Time", "Total", "Items Count", "Cashier"]
        )
        self.sales_detail_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.sales_detail_table)
        
        widget.setLayout(layout)
        return widget

    def create_top_items_tab(self):
        """Create top selling items tab."""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Controls
        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Show top:"))
        self.top_items_limit = QSpinBox()
        self.top_items_limit.setValue(10)
        self.top_items_limit.setMinimum(1)
        self.top_items_limit.setMaximum(100)
        control_layout.addWidget(self.top_items_limit)
        
        self.top_items_refresh_btn = QPushButton("Refresh")
        self.top_items_refresh_btn.setObjectName("search_button")
        control_layout.addWidget(self.top_items_refresh_btn)
        self.top_items_export_btn = QPushButton("Export to PDF")
        self.top_items_export_btn.setObjectName("secondary_button")
        control_layout.addWidget(self.top_items_export_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Top items table
        self.top_items_table = QTableWidget()
        self.top_items_table.setColumnCount(6)
        self.top_items_table.setHorizontalHeaderLabels(
            ["Item ID", "Part Name", "Category", "Quantity Sold", "Total Revenue", "Avg Price"]
        )
        self.top_items_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.top_items_table)
        
        widget.setLayout(layout)
        return widget

    def create_inventory_tab(self):
        """Create inventory status tab."""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Controls
        control_layout = QHBoxLayout()
        self.inventory_refresh_btn = QPushButton("Refresh")
        self.inventory_refresh_btn.setObjectName("search_button")
        control_layout.addWidget(self.inventory_refresh_btn)
        self.inventory_export_btn = QPushButton("Export to PDF")
        self.inventory_export_btn.setObjectName("secondary_button")
        control_layout.addWidget(self.inventory_export_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Inventory table
        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(9)
        self.inventory_table.setHorizontalHeaderLabels(
            ["ID", "Part Name", "Category", "Quantity", "Cost Price", "Selling Price", 
             "Total Cost Value", "Total Selling Value", "Potential Profit"]
        )
        self.inventory_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.inventory_table)
        
        widget.setLayout(layout)
        return widget

    def create_low_stock_tab(self):
        """Create low stock alert tab."""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Controls
        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Stock threshold:"))
        self.low_stock_threshold = QSpinBox()
        self.low_stock_threshold.setValue(5)
        self.low_stock_threshold.setMinimum(1)
        self.low_stock_threshold.setMaximum(100)
        control_layout.addWidget(self.low_stock_threshold)
        
        self.low_stock_refresh_btn = QPushButton("Refresh")
        self.low_stock_refresh_btn.setObjectName("search_button")
        control_layout.addWidget(self.low_stock_refresh_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Low stock table
        self.low_stock_table = QTableWidget()
        self.low_stock_table.setColumnCount(5)
        self.low_stock_table.setHorizontalHeaderLabels(
            ["ID", "Part Name", "Category", "Current Stock", "Supplier ID"]
        )
        self.low_stock_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.low_stock_table)
        
        widget.setLayout(layout)
        return widget

    def create_supplier_tab(self):
        """Create supplier performance tab."""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Controls
        control_layout = QHBoxLayout()
        self.supplier_refresh_btn = QPushButton("Refresh")
        self.supplier_refresh_btn.setObjectName("search_button")
        control_layout.addWidget(self.supplier_refresh_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Supplier table
        self.supplier_table = QTableWidget()
        self.supplier_table.setColumnCount(5)
        self.supplier_table.setHorizontalHeaderLabels(
            ["Supplier ID", "Supplier Name", "Number of Sales", "Total Items Sold", "Total Revenue"]
        )
        self.supplier_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.supplier_table)
        
        widget.setLayout(layout)
        return widget

    def create_category_tab(self):
        """Create category performance tab."""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Controls
        control_layout = QHBoxLayout()
        self.category_refresh_btn = QPushButton("Refresh")
        self.category_refresh_btn.setObjectName("search_button")
        control_layout.addWidget(self.category_refresh_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Category table
        self.category_table = QTableWidget()
        self.category_table.setColumnCount(5)
        self.category_table.setHorizontalHeaderLabels(
            ["Category", "Number of Sales", "Total Quantity", "Total Revenue", "Avg Price"]
        )
        self.category_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.category_table)
        
        widget.setLayout(layout)
        return widget

    def populate_sales_by_date_table(self, data):
        """Populate sales by date table."""
        self.sales_by_date_table.setRowCount(len(data))
        for row, item in enumerate(data):
            self.sales_by_date_table.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.sales_by_date_table.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.sales_by_date_table.setItem(row, 2, QTableWidgetItem(f"Php {item[2]:,.2f}" if item[2] else "Php 0.00"))
            self.sales_by_date_table.setItem(row, 3, QTableWidgetItem(f"Php {item[3]:,.2f}" if item[3] else "Php 0.00"))

    def populate_sales_detail_table(self, data):
        """Populate detailed transactions table."""
        self.sales_detail_table.setRowCount(len(data))
        for row, item in enumerate(data):
            # item indices: 0=id,1=date,2=time,3=total,4=items_count,5=cashier
            self.sales_detail_table.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.sales_detail_table.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.sales_detail_table.setItem(row, 2, QTableWidgetItem(str(item[2])))
            self.sales_detail_table.setItem(row, 3, QTableWidgetItem(f"Php {item[3]:,.2f}"))
            self.sales_detail_table.setItem(row, 4, QTableWidgetItem(str(item[4])))
            self.sales_detail_table.setItem(row, 5, QTableWidgetItem(str(item[5]) if item[5] else "Unknown"))

    def populate_top_items_table(self, data):
        """Populate top items table."""
        self.top_items_table.setRowCount(len(data))
        for row, item in enumerate(data):
            self.top_items_table.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.top_items_table.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.top_items_table.setItem(row, 2, QTableWidgetItem(str(item[2])))
            self.top_items_table.setItem(row, 3, QTableWidgetItem(str(item[3])))
            self.top_items_table.setItem(row, 4, QTableWidgetItem(f"Php {item[4]:,.2f}" if item[4] else "Php 0.00"))
            self.top_items_table.setItem(row, 5, QTableWidgetItem(f"Php {item[5]:,.2f}" if item[5] else "Php 0.00"))

    def populate_inventory_table(self, data):
        """Populate inventory table."""
        self.inventory_table.setRowCount(len(data))
        for row, item in enumerate(data):
            self.inventory_table.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.inventory_table.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.inventory_table.setItem(row, 2, QTableWidgetItem(str(item[2])))
            self.inventory_table.setItem(row, 3, QTableWidgetItem(str(item[3])))
            self.inventory_table.setItem(row, 4, QTableWidgetItem(f"Php {item[4]:,.2f}"))
            self.inventory_table.setItem(row, 5, QTableWidgetItem(f"Php {item[5]:,.2f}"))
            self.inventory_table.setItem(row, 6, QTableWidgetItem(f"Php {item[6]:,.2f}"))
            self.inventory_table.setItem(row, 7, QTableWidgetItem(f"Php {item[7]:,.2f}"))
            self.inventory_table.setItem(row, 8, QTableWidgetItem(f"Php {item[8]:,.2f}"))

    def populate_low_stock_table(self, data):
        """Populate low stock table."""
        self.low_stock_table.setRowCount(len(data))
        for row, item in enumerate(data):
            self.low_stock_table.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.low_stock_table.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.low_stock_table.setItem(row, 2, QTableWidgetItem(str(item[2])))
            self.low_stock_table.setItem(row, 3, QTableWidgetItem(str(item[3])))
            self.low_stock_table.setItem(row, 4, QTableWidgetItem(str(item[4])))

    def populate_supplier_table(self, data):
        """Populate supplier table."""
        self.supplier_table.setRowCount(len(data))
        for row, item in enumerate(data):
            self.supplier_table.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.supplier_table.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.supplier_table.setItem(row, 2, QTableWidgetItem(str(item[2]) if item[2] else "0"))
            self.supplier_table.setItem(row, 3, QTableWidgetItem(str(item[3]) if item[3] else "0"))
            self.supplier_table.setItem(row, 4, QTableWidgetItem(f"Php {item[4]:,.2f}" if item[4] else "Php 0.00"))

    def populate_category_table(self, data):
        """Populate category table."""
        self.category_table.setRowCount(len(data))
        for row, item in enumerate(data):
            self.category_table.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.category_table.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.category_table.setItem(row, 2, QTableWidgetItem(str(item[2])))
            self.category_table.setItem(row, 3, QTableWidgetItem(f"Php {item[3]:,.2f}" if item[3] else "Php 0.00"))
            self.category_table.setItem(row, 4, QTableWidgetItem(f"Php {item[4]:,.2f}" if item[4] else "Php 0.00"))

    def _export_table_widget_to_pdf(self, table_widget, filename, title=None, filters=None, metrics=None):
        """Export a table widget to PDF with professional formatting."""
        try:
            # Create professional report generator
            report = ProfessionalPDFReportGenerator(filename, landscape_mode=True)
            
            # Add header and metadata
            report.add_header()
            report.add_title_and_metadata(title or "Report", filters=filters, metrics=metrics)
            
            # Build data rows
            headers = [
                table_widget.horizontalHeaderItem(c).text()
                if table_widget.horizontalHeaderItem(c)
                else ""
                for c in range(table_widget.columnCount())
            ]

            data = []
            for r in range(table_widget.rowCount()):
                row = []
                for c in range(table_widget.columnCount()):
                    item = table_widget.item(r, c)
                    text = item.text() if item else ""
                    row.append(Paragraph(text, getSampleStyleSheet()['Normal']))
                data.append(row)
            
            # Add data table
            report.add_data_table(None, headers, data)
            
            # Add footer note
            report.add_footer_note("This is a confidential business report. Protect accordingly.")
            
            # Build PDF
            report.build()
            QMessageBox.information(self, "Success", f"Exported PDF to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export PDF: {str(e)}")

    def connect_export_buttons(self):
        # Connect export buttons to the handlers (call after controller sets up view)
        try:
            self.sales_export_pdf_btn.clicked.connect(self._handle_export_sales_pdf)
        except Exception:
            pass
        try:
            self.top_items_export_btn.clicked.connect(self._handle_export_top_items_pdf)
        except Exception:
            pass
        try:
            self.inventory_export_btn.clicked.connect(self._handle_export_inventory_pdf)
        except Exception:
            pass

    def _choose_save_file(self, suggested_name):
        path, _ = QFileDialog.getSaveFileName(self, "Save PDF", suggested_name, "PDF Files (*.pdf)")
        return path

    def _handle_export_sales_pdf(self):
        filename = self._choose_save_file("sales_report.pdf")
        if not filename:
            return
        # Get date range for filters
        start_date = self.sales_start_date.date().toString("yyyy-MM-dd")
        end_date = self.sales_end_date.date().toString("yyyy-MM-dd")
        
        filters = {
            "Date Range": f"{start_date} to {end_date}",
            "Report Type": "Detailed Transactions"
        }
        
        # Calculate metrics from sales summary table
        total_revenue = 0.0
        total_transactions = 0
        
        for r in range(self.sales_by_date_table.rowCount()):
            try:
                transactions_item = self.sales_by_date_table.item(r, 1)
                revenue_item = self.sales_by_date_table.item(r, 2)
                if transactions_item:
                    total_transactions += int(transactions_item.text())
                if revenue_item:
                    revenue_text = revenue_item.text().replace('Php ', '').replace(',', '')
                    total_revenue += float(revenue_text)
            except Exception:
                pass
        
        metrics = {
            "Total Records": str(self.sales_detail_table.rowCount()),
            "Total Revenue": f"Php {total_revenue:,.2f}",
            "Total Transactions": str(total_transactions),
        }
        
        title = "Sales Detailed Transactions Report"
        self._export_table_widget_to_pdf(self.sales_detail_table, filename, title=title, filters=filters, metrics=metrics)

    def _handle_export_top_items_pdf(self):
        filename = self._choose_save_file("top_items_report.pdf")
        if not filename:
            return
        
        limit = self.top_items_limit.value()
        filters = {
            "Top Items Limit": str(limit)
        }
        
        # Calculate metrics from top items table
        total_qty = 0
        total_revenue = 0.0
        
        for r in range(self.top_items_table.rowCount()):
            try:
                qty_item = self.top_items_table.item(r, 3)
                revenue_item = self.top_items_table.item(r, 4)
                if qty_item:
                    total_qty += int(qty_item.text())
                if revenue_item:
                    revenue_text = revenue_item.text().replace('Php ', '').replace(',', '')
                    total_revenue += float(revenue_text)
            except Exception:
                pass
        
        metrics = {
            "Items Shown": str(self.top_items_table.rowCount()),
            "Total Quantity": str(total_qty),
            "Total Revenue": f"Php {total_revenue:,.2f}",
        }
        
        title = "Top Selling Items Report"
        self._export_table_widget_to_pdf(self.top_items_table, filename, title=title, filters=filters, metrics=metrics)

    def _handle_export_inventory_pdf(self):
        filename = self._choose_save_file("inventory_report.pdf")
        if not filename:
            return
        
        # Calculate metrics from inventory table
        total_quantity = 0
        total_cost_value = 0.0
        total_selling_value = 0.0
        
        for r in range(self.inventory_table.rowCount()):
            try:
                qty_item = self.inventory_table.item(r, 3)
                cost_item = self.inventory_table.item(r, 6)
                selling_item = self.inventory_table.item(r, 7)
                
                if qty_item:
                    total_quantity += int(qty_item.text())
                if cost_item:
                    cost_text = cost_item.text().replace('Php ', '').replace(',', '')
                    total_cost_value += float(cost_text)
                if selling_item:
                    selling_text = selling_item.text().replace('Php ', '').replace(',', '')
                    total_selling_value += float(selling_text)
            except Exception:
                pass
        
        metrics = {
            "Total Items": str(self.inventory_table.rowCount()),
            "Total Quantity": str(total_quantity),
            "Total Inventory Value": f"Php {total_selling_value:,.2f}",
        }
        
        title = "Inventory Status Report"
        self._export_table_widget_to_pdf(self.inventory_table, filename, title=title, metrics=metrics)

    def export_to_csv(self, data, filename):
        """Export data to CSV file."""
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                # Write headers
                headers = ["Transaction ID", "Sale Date", "Part Name", "Quantity", "Unit Price", "Line Total"]
                writer.writerow(headers)
                # Write data
                for row in data:
                    writer.writerow(row)
            QMessageBox.information(None, "Success", f"Report exported to {filename}")
        except Exception as e:
            QMessageBox.critical(None, "Error", f"Failed to export: {str(e)}")

    def export_to_xlsx(self, data, filename):
        """Export data to XLSX (Excel) with proper types and column widths."""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Dependency Missing", "openpyxl is not installed. Please install openpyxl and try again.")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Report"

            headers = ["Transaction ID", "Sale Date", "Transaction Total", "Cashier", "Part Name", "Quantity", "Unit Price", "Line Total"]
            ws.append(headers)

     
            col_widths = [len(h) for h in headers]

            for record in data:
           
                tx_id = record[0]
                sale_dt = record[1]
                tx_total = record[2]
                cashier = record[3]
                part_name = record[4]
                qty = record[5]
                unit_price = record[6]
                line_total = record[7]

                # Try to coerce sale_dt to a datetime if it's a string
                if isinstance(sale_dt, str):
                    try:
                        # Try common formats
                        sale_dt_val = datetime.fromisoformat(sale_dt)
                    except Exception:
                        try:
                            sale_dt_val = datetime.strptime(sale_dt, "%Y-%m-%d %H:%M:%S")
                        except Exception:
                            sale_dt_val = sale_dt
                else:
                    sale_dt_val = sale_dt

                row = [tx_id, sale_dt_val, tx_total, cashier, part_name, qty, unit_price, line_total]
                ws.append(row)

                # update widths
                for i, value in enumerate(row):
                    text = str(value) if value is not None else ""
                    if len(text) > col_widths[i]:
                        col_widths[i] = len(text)

            # Apply formatting for numeric columns and date column
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
                # Sale Date is column 2
                cell_date = row[1]
                try:
                    if isinstance(cell_date.value, datetime):
                        cell_date.number_format = 'yyyy-mm-dd hh:mm:ss'
                except Exception:
                    pass

            # Apply number format for currency/number columns: Transaction Total (3), Unit Price (7), Line Total (8) using 1-based index
            for r in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
                try:
                    # transaction total at index 3 (0-based 2)
                    if isinstance(r[2].value, (int, float)):
                        r[2].number_format = FORMAT_NUMBER_00
                    # quantity at index 5
                    if isinstance(r[5].value, (int, float)):
                        r[5].number_format = FORMAT_NUMBER_00
                    # unit price index 6
                    if isinstance(r[6].value, (int, float)):
                        r[6].number_format = FORMAT_NUMBER_00
                    # line total index 7
                    if isinstance(r[7].value, (int, float)):
                        r[7].number_format = FORMAT_NUMBER_00
                except Exception:
                    pass

            # Set column widths
            for i, width in enumerate(col_widths, start=1):
                col_letter = get_column_letter(i)
                # Set a minimum width and add some padding
                ws.column_dimensions[col_letter].width = max(10, width + 2)

            # Align headers center
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center')

            wb.save(filename)
            QMessageBox.information(self, "Success", f"Report exported to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export XLSX: {str(e)}")
